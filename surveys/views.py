
import re
import datetime
from django.utils.timezone import localtime

from django.db.models import Q


from django.core.exceptions import ValidationError
from django.core.validators import validate_email

from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.contrib.auth.hashers import make_password

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

from django.http import FileResponse

from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

from django.views.generic import TemplateView, ListView
from django.contrib.auth.views import LoginView, LogoutView

from .models import Survey, Rater, Response, Choice, User, Client, Cohort
from .forms import SurveyPageForm, UploadFileForm, SurveySettingsForm

from django.urls import reverse, reverse_lazy

class IndexView(TemplateView):
    template_name = "surveys/index.html"

class LoginView(LoginView):
    template_name = 'surveys/login.html'

    def get_success_url(self):
        return reverse_lazy('surveys:surveys')

class LogoutView(LogoutView):
    template_name = 'surveys/logout.html'

class ThanksView(TemplateView):
    template_name = "surveys/thanks.html"

def raters_view(request, survey_id, rater_id):

    survey = get_object_or_404(Survey, id=survey_id)
    rater = get_object_or_404(Rater, id=rater_id)
    raters_list = survey.raters.all().filter(ratee_user=request.user.id).order_by('survey_progress', '-type')

    context = {'survey': survey, 'rater': rater, 'raters_list': raters_list}
    return render(request, "surveys/raters.html", context)


class SurveysView(LoginRequiredMixin, ListView):
    model = Survey
    template_name = "surveys/surveys.html"
    
    def get_queryset(self):
        return Survey.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["surveys"] = self.get_surveys_data()
        return context
    
    def get_surveys_data(self):
        surveys = self.get_queryset()
        data = {}     
        for survey in surveys:
            if survey.raters.filter(rater_user_id=self.request.user.id) != []:
                for counter, rater in enumerate(survey.raters.filter(rater_user_id=self.request.user.id)):
                    data[f"{survey.id}_{counter}"] = {
                        'survey': survey,
                        'rater': rater,
                    }
        return data

def navbar_view(request):

    superuser = [
        ("Surveys", reverse_lazy('surveys:surveys')),
        ("Log out", reverse_lazy('surveys:logout')),
        ]
    
    user = [
        ("Surveys", reverse_lazy('surveys:surveys')),
        ("Log out", reverse_lazy('surveys:logout')),
        ]
    
    login = [
        ("Log in", reverse_lazy('surveys:login')),
        ]
    
    if request.user.is_authenticated:
        if request.user.is_superuser:
            menu = superuser
        else:
            menu = user
    else:
        menu = login

    context = {'menu': menu}
    return render(request, "surveys/navbar.html", context)

def page_view(request, survey_id, rater_id):

    survey = get_object_or_404(Survey, id=survey_id)
    rater = get_object_or_404(Rater, id=rater_id)
     
    def save_response(data_items, form_valid_and_next=False):
        for field, value in data_items:
            if re.search('(radio_)\d+', field):
                if form_valid_and_next and not value == "":
                    data = None
                else:
                    data = {
                        'rater_id': rater_id,                     
                        'question_id': field[6:],
                        'choice_id': None if value == "" else value,
                        'multichoices': None,
                        'text': "",
                    }
            elif re.search('(multi_)\d+', field):
                if form_valid_and_next and not value == []:
                    data = None
                else:
                    multichoices = Choice.objects.all().filter(pk__in=request.POST.getlist(field))
                    data = {
                        'rater_id': rater_id,                     
                        'question_id': field[6:],
                        'multichoices': None if value == [] else multichoices,
                        'choice_id': None,
                        'text': "",
                        } 
            elif re.search('(open_)\d+', field):
                if form_valid_and_next and value == "":
                    data = None
                else:
                    data = {
                        'rater_id': rater_id,                     
                        'question_id': field[5:],
                        'text': value,
                        'multichoices': None,
                        'choice_id': None,
                    }
            else:
                data = None
            if data:
                try:
                    response = Response.objects.get(
                        rater_id=data['rater_id'],
                        question_id=data ['question_id']
                        )
                    response.choice_id = data['choice_id']
                    response.text = data['text']
                    response.save()
                    if data['multichoices']:
                        response.multichoices.set(data['multichoices'])
                    else:
                        response.multichoices.clear()
                except Response.DoesNotExist:
                    response = Response(
                        rater_id=data['rater_id'],
                        question_id=data ['question_id'],
                        choice_id=data['choice_id'],
                        text=data['text'])
                    response.save()
                    if data['multichoices']:
                        response.multichoices.set(data['multichoices'])
                    else:
                        response.multichoices.clear()
                if rater.survey_progress == 'unstarted':
                    rater.survey_progress = 'incomplete'
                    rater.save()
    
    if request.method == 'GET':
        if rater.survey_progress == "finished":
            raise Http404
        if request.method == 'GET':
            if 'back' in request.GET:
                if rater.survey_page_number > 1:
                    rater.survey_page_number -= 1
                    rater.save()
    
    if request.method == 'POST':
        form = SurveyPageForm(survey, rater, request.POST)
        if form.is_valid() and 'next' in request.POST:
            save_response(form.cleaned_data.items(), form_valid_and_next=True)    
            if rater.survey_page_number == survey.pages.all().count():
                rater.survey_progress = 'finished'
                rater.survey_date_taken = datetime.datetime.now(datetime.timezone.utc)
                rater.save()
                return redirect('surveys:thanks')
            else:
                rater.survey_page_number += 1
                rater.save()
                form = SurveyPageForm(survey, rater)
        else:
            save_response(form.cleaned_data.items())
    else:
        form = SurveyPageForm(survey, rater)

    context = {'form': form, 'survey': survey, 'rater': rater}
    response = render(request, "surveys/partials/page.html", context)
    return response

def take_view(request, survey_id, rater_id):

    rater = get_object_or_404(Rater, id=rater_id)
    
    if request.method == 'GET':
        if rater.survey_progress == "finished":
            raise Http404
        
    context = {'survey_id': survey_id, 'rater_id': rater_id}
    return render(request, "surveys/take.html", context)

def survey_dashboard_view(request, survey_id):

    survey = get_object_or_404(Survey, id=survey_id)    
    raters = survey.raters.all().order_by('-ratee_user')
    users = User.objects.all()
    clients = Client.objects.all()
    cohorts = Cohort.objects.all()
    hx_trigger_deleted_raters = False
    errors_list = []

    if request.method == 'GET':
        if request.GET.get('deleted_rater_id'):
            responses = Response.objects.all()
            deleted_rater_id = request.GET.get('deleted_rater_id')
            deleted_rater = raters.get(id=deleted_rater_id)
            if deleted_rater.survey.multi_rater and deleted_rater.type == 'self':
                raters_to_delete = raters.filter(ratee_user__id=deleted_rater.rater_user.id)
                deleted_responses = responses.filter(rater_id__in=[rater.id for rater in raters_to_delete])
            else:
                raters_to_delete = deleted_rater
                deleted_responses = responses.filter(rater_id=deleted_rater_id)
            # raters_to_delete.delete()
            # deleted_responses.delete()
            # hx_trigger_deleted_raters = True

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():

            wb = load_workbook(request.FILES.get('file'))
            ws = wb.active

            columns = ['first name', 'last name', 'rater email', 'rater type', 'ratee email', 'client name', 'cohort name']
            raters_list = []
            for row in range(1, ws.max_row+1):
                rater_data = {}
                for column in range(0, 7):
                    cell = ws[row][column]
                    """validate top row column names and order"""
                    if row == 1:
                        if cell.value is None:
                            errors_list.append(f"Line: {row}. Cell: {cell.coordinate}. Error: empty column name.")
                        else:
                            if not cell.value.lower() == columns[column]:
                                errors_list.append(f"Line: {row}. Cell: {cell.coordinate}. Error: invalid column name or column order. Valid column names and order: {columns}.")
                    else:
                        """validate data types"""
                        if column == 0 or column == 1 or column == 5 or column == 6:                            
                            if type(cell.value) is not str and not cell.value is None:
                                errors_list.append(f"Line: {row}. Cell: {cell.coordinate}. Error: value can either be text or none.")
                        if column == 3:
                            if cell.value is None or type(cell.value) is not str and not cell.value.lower() in ['self', 'manager', 'peer', 'report']:
                                errors_list.append(f"Line: {row}. Cell: {cell.coordinate}. Error: value must be 'self', 'manager', 'peer' or 'report'.")
                        if column == 2 or column == 4:
                            if cell.value is None:
                                errors_list.append(f"Line: {row}. Cell: {cell.coordinate}. Error: email is required.")
                            else:
                                try:
                                    validate_email(cell.value)
                                except ValidationError:
                                    errors_list.append(f"Line: {row}. Cell: {cell.coordinate}. Error: {cell.value} is invalid email.")
                        rater_data[columns[column]] = cell.value
                if row > 1:
                    """validate existing raters and users"""
                    if rater_data['rater type'] == 'self' and rater_data['rater email'] != rater_data['ratee email']:
                        errors_list.append(f"Line: {row}. Error: rater type '{rater_data['rater type']}' must have the same rater and ratee emails.")
                    if raters.filter(
                        rater_user__email=rater_data['rater email'],
                        ratee_user__email=rater_data['ratee email'],
                        type=rater_data['rater type'],
                    ):
                        errors_list.append(f"Line: {row}. Error: Rater type '{rater_data['rater type']}' for {rater_data['ratee email']} already exists in database. Duplicates are now allowed.")
                    if raters.filter(
                        ratee_user__email=rater_data['ratee email'],
                        type='manager'
                    ):
                        errors_list.append(f"Line: {row}. Error: Rater type '{rater_data['rater type']}' for {rater_data['ratee email']} already exists in database. Only one rater type '{rater_data['rater type']}' allowed per a ratee.")
                    
                    
                    #checking added rater without ratee 
                    if raters.filter(
                        rater_user__email=rater_data['ratee email'],
                        type='manager'
                    ):
                        errors_list.append(f"Line: {row}. Error: Rater type '{rater_data['rater type']}' for {rater_data['ratee email']} already exists in database. Only one rater type '{rater_data['rater type']}' allowed per a ratee.")

                    

                    raters_list.append(rater_data)

            if errors_list == []:
                print(raters_list)








            new_user = User()
            new_rater = Rater
            new_client = Client
            new_cohort = Cohort        



                # try:
                #     raters.get(rater_user__email=data['rater_email'], ratee_user__email=data['ratee_email'], type=data['rater_type'])
                #     validation_errors.append(f"Duplicate {data['rater_email']} '{data['rater_type']}' in line {row}, cell {cell.coordinate} in db.")
                # except Rater.DoesNotExist:
                #     print(f"New rater: {data['rater_email']}")
                #     try:
                #         users.get(email=data['rater_email'])
                #     except User.DoesNotExist:
                #         print(f"New user: {data['rater_email']}")
                
            #     data['validation_errors'] = validation_errors
            #     valid_raters.append(data) if data['validation_errors'] == [] else unvalid_raters.append(data)
            # print(valid_raters)
            # print()
            # print(unvalid_raters)

            #         User.objects.create(
            #             email=rater['rater_email'],
            #             first_name = rater['first_name'],
            #             last_name = rater['last_name'],
            #             password=make_password(None),
            #             )
       
    else:
        form = UploadFileForm
    
        # if request.method == "POST":
    #     if 'invite_btn' in request.POST:
    #         send_mail(
    #             "Inviting you to take a survey!",
    #             "Just testing what goes in the body",
    #             "address_from",
    #             ["akokarpov@gmail.com"],
    #             fail_silently=False,
    #             html_message=f"""
    #             <h1>Hello, {rater.rater_name_or_email}!</h1>
    #             <p>Please <a href="http://127.0.0.1:8000{reverse('surveys:take', kwargs={'survey_id': survey_id, 'rater_id': rater_id})}">take {survey.name}</a> before {survey.end_date.strftime('%d.%m.%Y, %H:%m %Z')}.</p>
    #             {render_to_string("surveys/raters.html", {'survey': survey, 'rater': rater, 'raters_list': raters_list}, request)}
    #             """
    #             )
    
    context = {'raters': raters, 'survey': survey, 'form': form, 'errors_list': errors_list}
    response = render(request, "surveys/survey_dashboard.html", context)
    
    if hx_trigger_deleted_raters:
        response['HX-Trigger'] = 'deleted_raters'
    
    return response

def search_raters_view(request, survey_id):

    survey = get_object_or_404(Survey, id=survey_id)
    raters = survey.raters.all()
    
    raters = survey.raters.all()
    if request.method == 'GET':
        search = request.GET['search']
        if search:
            raters = raters.filter(
                Q(rater_user__first_name__icontains=search) |
                Q(rater_user__last_name__icontains=search) |
                Q(rater_user__email__icontains=search) |
                Q(type__icontains=search) |
                Q(ratee_user__email__icontains=search) |
                Q(cohort__name__icontains=search) |
                Q(cohort__client__name__icontains=search) |
                Q(survey_progress__icontains=search) |
                Q(survey_date_taken__icontains=search)
                )

    context = {'raters': raters, 'survey': survey}
    return render(request, "surveys/partials/raters_table.html", context)


def download_raters_view(request, survey_id):

    survey = get_object_or_404(Survey, id=survey_id)
    raters = survey.raters.all().order_by('-ratee_user')

    data = [
        [
            "First Name",
            "Last Name",
            "Rater Email",
            "Rater Type",
            "Ratee Email",
            "Client Name",
            "Cohort Name",
            "Survey Progress",
            "Survey Date Taken",
         ]
    ]

    for rater in raters:
        
        survey_date_taken = ""
        if rater.survey_date_taken:
            survey_date_taken = localtime(rater.survey_date_taken).replace(tzinfo=None).strftime("%d.%m.%Y")

        data.append(
            [
                rater.rater_user.first_name,
                rater.rater_user.last_name,
                rater.rater_user.email,
                rater.type,
                rater.ratee_user.email,
                rater.cohort.client.name,
                rater.cohort.name,
                rater.survey_progress,
                survey_date_taken,
            ]
        )
    
    fill_red = PatternFill(start_color='FCDCDC', end_color='FCDCDC', fill_type='solid')
    fill_green = PatternFill(start_color='B3FFBA', end_color='B3FFBA', fill_type='solid')
    fill_yellow = PatternFill(start_color='FDFFB3', end_color='FDFFB3', fill_type='solid')
    font_black_bold = Font(bold=True, color='000000')

    wb = Workbook()
    ws = wb.active
    
    for row in data:
        ws.append(row)
    
    #TODO: color required fields
    for row in ws["A1:I1"]:
        for cell in row:
            cell.font = font_black_bold
            cell.alignment = Alignment(horizontal='center')

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length
    
    rule1 = Rule(
        stopIfTrue=True,
        type='expression',
        formula=['$H1="unstarted"'],
        dxf=DifferentialStyle(fill=fill_red)
    )

    rule2 = Rule(
        stopIfTrue=True,
        type='expression',
        formula=['$H1="incomplete"'],
        dxf=DifferentialStyle(fill=fill_yellow)
    )

    rule3 = Rule(
        stopIfTrue=True,
        type='expression',
        formula=['$H1="finished"'],
        dxf=DifferentialStyle(fill=fill_green)
    )
    
    ws.conditional_formatting.add('A1:I100', rule1)
    ws.conditional_formatting.add('A1:I100', rule2)
    ws.conditional_formatting.add('A1:I100', rule3) 

    xls_stream = BytesIO()
    wb.save(xls_stream)
    xls_stream.seek(0)

    return FileResponse(xls_stream, as_attachment=True, filename='survey raters.xlsx')